# 1. IMPORTAMOS LAS NUEVAS LIBRERÍAS
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, flash, render_template
from flask_socketio import SocketIO, emit
import mysql.connector
from datetime import date, datetime, time
import datetime # Necesario para convertir la fecha a un formato enviable
from functools import wraps
import pandas as pd
from io import BytesIO
from flask import send_file
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import numpy as np
from fpdf import FPDF

# Configuración de la DB (sin cambios)
db_config = {
    'user': 'root',
    'password': '',
    'host': '127.0.0.1',
    'database': 'gestion_personal_db'
}

app = Flask(__name__)
# 2. CONFIGURACIÓN DE SOCKETIO
# Se necesita una 'secret key' para que funcionen las sesiones
app.config['SECRET_KEY'] = 'dariomayor' 
socketio = SocketIO(app)


# Filtro para formatear fechas de YYYY-MM-DD a DD/MM/YYYY en el HTML
def format_date_html_filter(date_str):
    try:
        # Asume que la entrada es un string YYYY-MM-DD
        return datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        return date_str # Devuelve el original si falla
    
app.jinja_env.filters['format_date_html'] = format_date_html_filter
    

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash("Por favor, inicia sesión para ver esta página.", "error")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function    

# --- 1. DECORADOR DE SEGURIDAD ---
# Esta función revisará si el usuario es "admin"
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Si no está logueado
        if 'logged_in' not in session:
            flash("Por favor, inicia sesión para ver esta página.", "error")
            return redirect(url_for('login'))
        # Si está logueado PERO NO es admin
        if session.get('role') != 'admin':
            flash("No tienes permiso de administrador para ver esta página.", "error")
            return redirect(url_for('dashboard')) # Lo manda al dashboard normal
        return f(*args, **kwargs)
    return decorated_function


# --- RUTA 1: El monitor visual ---
@app.route("/monitor")
def monitor():
    return render_template('monitor.html')

@app.route("/fichar_error", methods=['POST'])
def recibir_error_fichaje():
    """
    El ESP32 llama a esta ruta cuando el sensor no reconoce la huella.
    No toca la DB, solo avisa a los monitores.
    """
    print("--- Fichaje Fallido ---")
    print("Evento WebSocket 'huella_no_reconocida' emitido al monitor.")
    
    # Emite el nuevo evento a todos los monitores
    socketio.emit('huella_no_reconocida')
    
    # Responde OK al ESP32
    return jsonify({"status": "error_notificado"}), 200

# --- RUTA 2: El endpoint para el ESP32  ---
@app.route("/fichar", methods=['POST'])
@app.route("/fichar", methods=['POST'])
def recibir_fichaje():
    datos = request.json
    sensor_id = datos.get('id_huella') 
    
    if not sensor_id:
        return jsonify({"status": "error", "mensaje": "No se recibió id_huella"}), 400

    print(f"--- Nuevo Fichaje ---")
    print(f"Sensor ID recibido: {sensor_id}")

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True) 

        # PASO A: Buscar el usuario
        cursor.execute("SELECT usuario_id FROM huellas WHERE huella_id = %s", (sensor_id,))
        usuario_encontrado = cursor.fetchone() 

        if not usuario_encontrado:
            print(f"Error: Sensor ID {sensor_id} no está registrado.")
            return jsonify({"status": "error", "mensaje": "Huella no registrada"}), 404

        usuario_id = usuario_encontrado['usuario_id']
        print(f"Usuario encontrado: ID {usuario_id}")

        # PASO B: Decidir si es ENTRADA o SALIDA (Lógica Diaria)
        
        # 1. Buscamos si hay fichajes HOY para este usuario
        today = datetime.date.today()
        
        sql_ultimo_hoy = """
            SELECT tipo FROM fichajes 
            WHERE usuario_id = %s AND DATE(timestamp) = %s
            ORDER BY timestamp DESC 
            LIMIT 1
        """
        cursor.execute(sql_ultimo_hoy, (usuario_id, today))
        ultimo_fichaje_hoy = cursor.fetchone()

        nuevo_tipo = ""
        
        if not ultimo_fichaje_hoy:
            # CASO 1: No hay fichajes hoy.
            # Por lo tanto, es el primer movimiento del día -> ENTRADA
            # (Esto arregla automáticamente el olvido de salida de ayer)
            nuevo_tipo = "entrada"
            print(">> Primer fichaje del día: ENTRADA")
        else:
            # CASO 2: Ya fichó hoy. Seguimos la secuencia normal.
            tipo_anterior = ultimo_fichaje_hoy['tipo']
            
            if tipo_anterior == "entrada":
                nuevo_tipo = "salida"
            else: # Si el último fue salida (ej: salió a almorzar)
                nuevo_tipo = "entrada" # Vuelve a entrar
            
            print(f">> Fichaje previo hoy fue {tipo_anterior}. Nuevo: {nuevo_tipo}")
        
        # PASO C: Insertar el nuevo fichaje
        sql_insert = "INSERT INTO fichajes (usuario_id, timestamp, tipo) VALUES (%s, NOW(), %s)"
        valores = (usuario_id, nuevo_tipo)
        cursor.execute(sql_insert, valores)
        id_fichaje_nuevo = cursor.lastrowid
        conn.commit() 

        print(f"¡Éxito! Fichaje guardado en la DB.")
        
        # --- Notificar al Monitor ---
        sql_get_data = """
            SELECT f.id, f.timestamp, f.tipo, u.nombre, u.apellido 
            FROM fichajes AS f
            JOIN usuarios AS u ON f.usuario_id = u.id
            WHERE f.id = %s
        """
        cursor.execute(sql_get_data, (id_fichaje_nuevo,))
        datos_para_monitor = cursor.fetchone()
        
        if datos_para_monitor:
            datos_para_monitor['timestamp'] = datos_para_monitor['timestamp'].isoformat()
            socketio.emit('nuevo_fichaje', datos_para_monitor)
            print("Evento WebSocket emitido.")
            
        print("-----------------------\n")
        cursor.close()
        conn.close()

    except mysql.connector.Error as err:
        print(f"Error de Base de Datos: {err}")
        return jsonify({"status": "error", "mensaje": "Error de base de datos"}), 500

    return jsonify({"status": "ok", "recibido": sensor_id, "tipo": nuevo_tipo}), 200


# --- RUTA DE LOGIN  ---
@app.route("/login", methods=['GET', 'POST'])
def login():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # --- ¡ESTE ES EL ÚNICO CAMBIO! ---
        # Ahora busca en la tabla 'usuarios'
        cursor.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        # --- Fin del cambio ---
        
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        # El resto de la lógica es IDÉNTICA
        if user and user['password_hash'] and check_password_hash(user['password_hash'], password):
            session['logged_in'] = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['nombre'] = user['nombre']     
            session['apellido'] = user['apellido'] 
            print(f"Login exitoso para: {user['username']} (Rol: {user['role']})")
            return redirect(url_for('dashboard'))
        else:
            print("Fallo de login.")
            flash("Usuario o contraseña incorrectos", "error") 
            return redirect(url_for('login'))

    return render_template('login.html')


# --- RUTA DE LOGOUT ---
@app.route("/logout")
def logout():
    # Limpia la sesión
    session.clear()
    flash("Has cerrado sesión.", "success")
    return redirect(url_for('login'))


# --- RUTA DEL PANEL DE ADMINISTRACIÓN (PROTEGIDA) ---
@app.route("/dashboard")
def dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    # Preparamos las variables para pasarlas al HTML
    contexto = {
        'username': session['username'],
        'role': session['role'],
        'nombre': session['nombre'],       
        'apellido': session['apellido']  
    }

    return render_template('dashboard.html', **contexto)

@app.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Obtenemos todos los usuarios para la tabla
        cursor.execute("SELECT id, nombre, apellido, legajo, username, role FROM usuarios ORDER BY apellido, nombre")
        lista_de_usuarios = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('admin_usuarios.html', usuarios=lista_de_usuarios)
        
    except mysql.connector.Error as err:
        flash(f"Error al cargar usuarios: {err}", "error")
        return redirect(url_for('dashboard'))

# --- CREAR USUARIO ---
@app.route("/admin/crear", methods=['GET', 'POST'])
@admin_required
def crear_usuario():
    # 1. Obtenemos una copia de los datos del formulario (si el método es POST)
    form_data = request.form if request.method == 'POST' else {}

    if request.method == 'POST':
        # Recolección de datos
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        legajo = request.form['legajo']
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        huellas_str = request.form.get('huellas', '')
        
        password_hash = generate_password_hash(password)

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        try:
            # 1. Insertar el nuevo usuario
            sql_insert_user = """
                INSERT INTO usuarios (nombre, apellido, legajo, username, password_hash, role) 
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql_insert_user, (nombre, apellido, legajo, username, password_hash, role))
            new_user_id = cursor.lastrowid
            
            # 2. Insertar huellas
            if new_user_id and huellas_str:
                huellas_list = [h.strip() for h in huellas_str.split(',') if h.strip().isdigit()]
                if huellas_list:
                    sql_insert_huella = "INSERT INTO huellas (huella_id, usuario_id) VALUES (%s, %s)"
                    for huella_id in set(huellas_list):
                        cursor.execute(sql_insert_huella, (huella_id, new_user_id))
            
            conn.commit()
            flash(f"Usuario '{username}' y huellas vinculadas exitosamente.", "success")
            return redirect(url_for('admin_usuarios'))
        
        except mysql.connector.Error as err:
            conn.rollback()
            if err.errno == 1062: 
                # Si hay error, asignamos el error al mensaje flash
                flash(f"Error: El Email, Legajo o una de las Huellas ya está en uso. Por favor, revisa.", "error")
            else:
                flash(f"Error de base de datos: {err}", "error")
            
            # 2. Si hay error, renderizamos el formulario de nuevo 
            #    con todos los datos que el usuario introdujo (form_data)
            return render_template('crear_usuario.html', form_data=form_data, huellas_str=huellas_str)
        
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    # Si es GET, muestra el formulario por primera vez (form_data está vacío)
    return render_template('crear_usuario.html', form_data=form_data)

# --- EDITAR USUARIO ---
@app.route("/admin/editar/<int:usuario_id>", methods=['GET', 'POST'])
@admin_required
def editar_usuario(usuario_id):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    
    # 1. Si hay un POST fallido, los datos vienen del request.form
    # 2. Si es GET, primero los obtenemos de la DB
    usuario_a_editar = None
    huellas_str = ''
    
    if request.method == 'POST':
        # Los datos POST fallidos están en request.form
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        legajo = request.form['legajo']
        username = request.form['username']
        role = request.form['role']
        nueva_password = request.form['nueva_password']
        huellas_str = request.form.get('huellas', '')
        
        try:
            # Lógica de actualización (igual que antes)
            if nueva_password:
                password_hash = generate_password_hash(nueva_password)
                sql_update = """
                    UPDATE usuarios SET nombre=%s, apellido=%s, legajo=%s, username=%s, role=%s, password_hash=%s WHERE id=%s
                """
                cursor.execute(sql_update, (nombre, apellido, legajo, username, role, password_hash, usuario_id))
            else:
                sql_update = """
                    UPDATE usuarios SET nombre=%s, apellido=%s, legajo=%s, username=%s, role=%s WHERE id=%s
                """
                cursor.execute(sql_update, (nombre, apellido, legajo, username, role, usuario_id))
            
            # Sincronizar huellas
            cursor.execute("DELETE FROM huellas WHERE usuario_id = %s", (usuario_id,))
            if huellas_str:
                huellas_list = [h.strip() for h in huellas_str.split(',') if h.strip().isdigit()]
                if huellas_list:
                    sql_insert_huella = "INSERT INTO huellas (huella_id, usuario_id) VALUES (%s, %s)"
                    for huella_id in set(huellas_list):
                        cursor.execute(sql_insert_huella, (huella_id, usuario_id))
            
            conn.commit()
            flash(f"Usuario '{username}' y huellas actualizadas exitosamente.", "success")
            return redirect(url_for('admin_usuarios'))
        
        except mysql.connector.Error as err:
            conn.rollback() 
            if err.errno == 1062:
                flash(f"Error: El Email, Legajo o una de las Huellas ya está en uso.", "error")
            else:
                flash(f"Error de base de datos: {err}", "error")
            
            # Si hay error, creamos un diccionario con los datos fallidos 
            # para pasarlos a la plantilla.
            usuario_a_editar = {
                'id': usuario_id, 'nombre': nombre, 'apellido': apellido, 'legajo': legajo, 
                'username': username, 'role': role
            }
            # Y renderizamos el formulario de nuevo
            return render_template('editar_usuario.html', usuario=usuario_a_editar, huellas_str=huellas_str)
        
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    # Si es GET, obtenemos los datos de la DB para la primera carga
    try:
        cursor.execute("SELECT id, nombre, apellido, legajo, username, role FROM usuarios WHERE id = %s", (usuario_id,))
        usuario_a_editar = cursor.fetchone()
        
        cursor.execute("SELECT huella_id FROM huellas WHERE usuario_id = %s", (usuario_id,))
        huellas_raw = cursor.fetchall()
        huellas_str = ', '.join([str(h['huella_id']) for h in huellas_raw])
        
        if not usuario_a_editar:
            flash("Error: Usuario no encontrado.", "error")
            return redirect(url_for('dashboard'))
            
        return render_template('editar_usuario.html', usuario=usuario_a_editar, huellas_str=huellas_str)
    
    except mysql.connector.Error as err:
        flash(f"Error de base de datos: {err}", "error")
        return redirect(url_for('dashboard'))
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# --- ELIMINAR USUARIO ---
@app.route("/admin/eliminar/<int:usuario_id>", methods=['POST'])
@admin_required
def eliminar_usuario(usuario_id):
    # No puedes eliminarte a ti mismo
    if usuario_id == session['user_id']:
        flash("No puedes eliminar tu propia cuenta de administrador.", "error")
        return redirect(url_for('admin_usuarios'))

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # ¡IMPORTANTE! Borra los "hijos" primero para evitar errores
        # 1. Borra las huellas de ese usuario
        cursor.execute("DELETE FROM huellas WHERE usuario_id = %s", (usuario_id,))
        # 2. Borra los fichajes de ese usuario
        cursor.execute("DELETE FROM fichajes WHERE usuario_id = %s", (usuario_id,))
        # 3. Ahora sí, borra al usuario "padre"
        cursor.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
        
        conn.commit()
        flash("Usuario eliminado exitosamente (junto con sus huellas y fichajes asociados).", "success")
    except mysql.connector.Error as err:
        flash(f"Error de base de datos al eliminar: {err}", "error")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
    return redirect(url_for('admin_usuarios'))



@app.route("/admin/descargar_excel", methods=['POST'])
@admin_required
def descargar_excel():
    try:
        # 1. Obtener las fechas
        start_date_str = request.form['start_date']
        end_date_str = request.form['end_date']
        start_date = f"{start_date_str} 00:00:00"
        end_date = f"{end_date_str} 23:59:59"
        
        # --- OBTENER NOMBRE DEL MES ---
        try:
            start_date_obj = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
            meses = {
                1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
                7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
            }
            month_name = meses[start_date_obj.month].capitalize()
        except Exception:
            month_name = "" 
            
        print(f"Generando reporte HTML-Style Excel (Limpieza de duplicados) para {month_name}...")

        # 2. Conectar y ejecutar la consulta SQL
        conn = mysql.connector.connect(**db_config)
        sql_query = """
            SELECT 
                u.legajo, 
                u.nombre, 
                u.apellido, 
                f.timestamp, 
                f.tipo
            FROM fichajes f
            JOIN usuarios u ON f.usuario_id = u.id
            WHERE f.timestamp BETWEEN %s AND %s
            ORDER BY u.legajo, f.timestamp
        """
        df = pd.read_sql(sql_query, conn, params=(start_date, end_date))
        conn.close()

        if df.empty:
            flash("No se encontraron fichajes en ese rango de fechas.", "error")
            return redirect(url_for('dashboard'))

        # 3. PROCESAR DATOS (Lógica Min/Max)
        
        df['fecha'] = df['timestamp'].dt.date
        df['hora'] = df['timestamp'].dt.time
        
        entradas_df = df[df['tipo'] == 'entrada']
        salidas_df = df[df['tipo'] == 'salida']

        common_cols = ['legajo', 'nombre', 'apellido', 'fecha']
        
        # Calcular MIN (Turno Mañana)
        e_min = entradas_df.groupby(common_cols)['hora'].min().reset_index().rename(columns={'hora': 'Entrada'})
        s_min = salidas_df.groupby(common_cols)['hora'].min().reset_index().rename(columns={'hora': 'Salida'})
        
        # Calcular MAX (Turno Tarde)
        e_max = entradas_df.groupby(common_cols)['hora'].max().reset_index().rename(columns={'hora': 'Entrada'})
        s_max = salidas_df.groupby(common_cols)['hora'].max().reset_index().rename(columns={'hora': 'Salida'})

        # Limpieza de duplicados: Si el MAX es igual al MIN, lo borramos del MAX.
        # 1. Limpiar Entradas repetidas
        e_max = pd.merge(e_max, e_min, on=common_cols, suffixes=('', '_min'), how='left')
        mask_e = e_max['Entrada'] == e_max['Entrada_min']
        e_max.loc[mask_e, 'Entrada'] = None # Dejamos vacío si es igual al de la mañana
        e_max = e_max.drop(columns=['Entrada_min']) # Limpiamos columna auxiliar
        
        # 2. Limpiar Salidas repetidas (Lo que pediste específicamente)
        s_max = pd.merge(s_max, s_min, on=common_cols, suffixes=('', '_min'), how='left')
        mask_s = s_max['Salida'] == s_max['Salida_min']
        s_max.loc[mask_s, 'Salida'] = None # Dejamos vacío si es igual a la salida de la mañana
        s_max = s_max.drop(columns=['Salida_min']) # Limpiamos columna auxiliar
        # -----------------------------------

        e_min['Turno'] = 'Mañana'
        e_max['Turno'] = 'Tarde'
        s_min['Turno'] = 'Mañana'
        s_max['Turno'] = 'Tarde'

        all_entradas = pd.concat([e_min, e_max])
        all_salidas = pd.concat([s_min, s_max])

        merge_cols = ['legajo', 'nombre', 'apellido', 'fecha', 'Turno']
        reporte_long_df = pd.merge(all_entradas, all_salidas, on=merge_cols, how='outer')

        # Limpieza extra por si quedaron duplicados vacíos
        reporte_long_df = reporte_long_df.drop_duplicates(subset=['legajo', 'fecha', 'Entrada', 'Salida'])

        reporte_long_df['Apellido y Nombre'] = reporte_long_df['apellido'].fillna('') + ', ' + reporte_long_df['nombre'].fillna('')
        
        reporte_long_df['Entrada'] = reporte_long_df['Entrada'].apply(lambda x: x.strftime('%H:%M') if isinstance(x, datetime.time) else '')
        reporte_long_df['Salida'] = reporte_long_df['Salida'].apply(lambda x: x.strftime('%H:%M') if isinstance(x, datetime.time) else '')

        pivot_df = reporte_long_df.pivot_table(
            index=['legajo', 'Apellido y Nombre', 'Turno'],
            columns='fecha',
            values=['Entrada', 'Salida'],
            aggfunc='first',
            fill_value=''
        )

        if pivot_df.empty:
             flash("No se generaron datos pivotados.", "error")
             return redirect(url_for('dashboard'))

        pivot_df = pivot_df.swaplevel(0, 1, axis=1).sort_index(axis=1)
        pivot_df = pivot_df.reindex(pd.Categorical(['Mañana', 'Tarde'], ordered=True), level='Turno')
        pivot_df = pivot_df.sort_index(level=['legajo', 'Turno'])
        pivot_df = pivot_df.reset_index()

        # 4. Crear el archivo Excel 
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        workbook = writer.book
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        worksheet = workbook.create_sheet(title='Reporte Fichajes', index=0)
        writer.sheets['Reporte Fichajes'] = worksheet

        # 5. Definir Estilos (Igual que antes)
        bold_font = Font(bold=True)
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        
        medium_side = Side(style='medium') 
        thin_side = Side(style='thin')
        gray_thin_side = Side(style='thin', color="A9A9A9") 
        
        border_R1 = Border(top=medium_side, left=medium_side, right=medium_side, bottom=thin_side)
        border_A2 = Border(top=thin_side, left=medium_side, bottom=medium_side, right=thin_side)
        border_B2 = Border(top=thin_side, left=thin_side, bottom=medium_side, right=medium_side)
        border_R2_entrada = Border(top=thin_side, left=medium_side, bottom=medium_side, right=thin_side)
        border_R2_salida = Border(top=thin_side, left=thin_side, bottom=medium_side, right=medium_side)
        
        # Bordes de Datos
        border_legajo_top = Border(left=medium_side, top=medium_side, right=thin_side, bottom=thin_side)
        border_legajo_bottom = Border(left=medium_side, top=thin_side, right=thin_side, bottom=medium_side)
        border_nombre_top = Border(left=thin_side, top=medium_side, right=medium_side, bottom=thin_side)
        border_nombre_bottom = Border(left=thin_side, top=thin_side, right=medium_side, bottom=medium_side)

        border_data_entrada_am = Border(left=medium_side, top=medium_side, right=gray_thin_side, bottom=gray_thin_side) 
        border_data_salida_am = Border(left=gray_thin_side, top=medium_side, right=medium_side, bottom=gray_thin_side)
        border_data_entrada_pm = Border(left=medium_side, top=gray_thin_side, right=gray_thin_side, bottom=medium_side) 
        border_data_salida_pm = Border(left=gray_thin_side, top=gray_thin_side, right=medium_side, bottom=medium_side)

        # 6. Escribir Encabezado Fila 1 
        cell_a1 = worksheet['A1']
        cell_a1.value = f"Mar del Plata - {month_name}"
        cell_a1.font = bold_font
        cell_a1.fill = silver_fill
        cell_a1.border = border_R1
        cell_a1.alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:B1')
        cell_b1 = worksheet['B1']
        cell_b1.font = bold_font
        cell_b1.fill = silver_fill
        cell_b1.border = border_R1
        cell_b1.alignment = Alignment(horizontal='center')
        
        fechas_unicas = sorted(list(set([col[0] for col in pivot_df.columns if isinstance(col[0], datetime.date)])))
        
        col_idx = 3 
        for fecha in fechas_unicas:
            worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = fecha.strftime('%d/%m/%Y')
            cell.alignment = Alignment(horizontal='center')
            cell.font = bold_font
            cell.fill = silver_fill
            cell.border = border_R1
            cell_plus_1 = worksheet.cell(row=1, column=col_idx + 1)
            cell_plus_1.font = bold_font
            cell_plus_1.fill = silver_fill
            cell_plus_1.border = border_R1
            col_idx += 2

        # 7. Escribir Encabezado Fila 2
        cell_a2 = worksheet['A2']
        cell_a2.value = "Legajo"
        cell_a2.font = bold_font
        cell_a2.fill = silver_fill
        cell_a2.border = border_A2
        cell_a2.alignment = Alignment(horizontal='center')
        
        cell_b2 = worksheet['B2']
        cell_b2.value = "Apellido y Nombre"
        cell_b2.font = bold_font
        cell_b2.fill = silver_fill
        cell_b2.border = border_B2
        cell_b2.alignment = Alignment(horizontal='center')
        
        col_idx = 3 
        for fecha in fechas_unicas:
            cell_e = worksheet.cell(row=2, column=col_idx)
            cell_e.value = "Entrada"
            cell_e.font = bold_font
            cell_e.fill = silver_fill
            cell_e.border = border_R2_entrada
            cell_e.alignment = Alignment(horizontal='center')
            cell_s = worksheet.cell(row=2, column=col_idx + 1)
            cell_s.value = "Salida"
            cell_s.font = bold_font
            cell_s.fill = silver_fill
            cell_s.border = border_R2_salida
            cell_s.alignment = Alignment(horizontal='center')
            col_idx += 2
        
        # 8. Escribir datos con fusión y bordes de bloque
        current_excel_row = 3
        
        for (legajo, apellido_nombre), group in pivot_df.groupby(['legajo', 'Apellido y Nombre']):
            
            row_am = group[group['Turno'] == 'Mañana'] 
            row_pm = group[group['Turno'] == 'Tarde']   
            
            # --- Escribir y Fusionar (Rowspan) ---
            cell_legajo = worksheet.cell(row=current_excel_row, column=1, value=legajo)
            cell_legajo.border = border_legajo_top
            cell_legajo.alignment = Alignment(vertical='center', horizontal='center')
            worksheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row + 1, end_column=1)
            worksheet.cell(row=current_excel_row+1, column=1).border = border_legajo_bottom

            cell_nombre = worksheet.cell(row=current_excel_row, column=2, value=apellido_nombre)
            cell_nombre.border = border_nombre_top
            cell_nombre.alignment = Alignment(vertical='center')
            worksheet.merge_cells(start_row=current_excel_row, start_column=2, end_row=current_excel_row + 1, end_column=2)
            worksheet.cell(row=current_excel_row+1, column=2).border = border_nombre_bottom

            # --- Escribir datos (Fila 1 del empleado) ---
            col_idx = 3
            for fecha in fechas_unicas:
                val_e = row_am[(fecha, 'Entrada')].values[0] if not row_am.empty else ''
                cell_e_am = worksheet.cell(row=current_excel_row, column=col_idx, value=val_e)
                cell_e_am.border = border_data_entrada_am
                cell_e_am.alignment = Alignment(horizontal='center')
                
                val_s = row_am[(fecha, 'Salida')].values[0] if not row_am.empty else ''
                cell_s_am = worksheet.cell(row=current_excel_row, column=col_idx + 1, value=val_s)
                cell_s_am.border = border_data_salida_am
                cell_s_am.alignment = Alignment(horizontal='center')
                
                col_idx += 2
            
            current_excel_row += 1
            
            # --- Escribir datos (Fila 2 del empleado) ---
            col_idx = 3
            for fecha in fechas_unicas:
                val_e = row_pm[(fecha, 'Entrada')].values[0] if not row_pm.empty else ''
                cell_e_pm = worksheet.cell(row=current_excel_row, column=col_idx, value=val_e)
                cell_e_pm.border = border_data_entrada_pm
                cell_e_pm.alignment = Alignment(horizontal='center')

                val_s = row_pm[(fecha, 'Salida')].values[0] if not row_pm.empty else ''
                cell_s_pm = worksheet.cell(row=current_excel_row, column=col_idx + 1, value=val_s)
                cell_s_pm.border = border_data_salida_pm
                cell_s_pm.alignment = Alignment(horizontal='center')
                
                col_idx += 2
                
            current_excel_row += 1

        # 9. Ajustar anchos
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['A'].width = 15

        # 10. Guardar y enviar
        writer.close() 
        output.seek(0) 

        print("Reporte Excel estilo HTML (Limpio) generado. Enviando al usuario...")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_fichaje.xlsx'
        )

    except Exception as e:
        import traceback
        print("--- ERROR DETALLADO AL GENERAR EXCEL ---")
        traceback.print_exc()
        print("-----------------------------------------")
        
        flash(f"Error al generar el reporte: {e}", "error")
        return redirect(url_for('dashboard'))
    
    
# --- RUTA PARA VER EL LOG DE FICHAJES ---
@app.route("/ver_fichajes", methods=['POST'])
@login_required
def ver_fichajes():
    try:
        start_date_str = request.form['log_start_date']
        end_date_str = request.form['log_end_date']
        
        start_date = f"{start_date_str} 00:00:00"
        end_date = f"{end_date_str} 23:59:59"

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # --- LÓGICA DE FILTRO POR ROL (Se queda aquí) ---
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        sql_log = """
            SELECT 
                f.timestamp, f.tipo, u.nombre, u.apellido, u.legajo
            FROM fichajes f
            JOIN usuarios u ON f.usuario_id = u.id
            WHERE f.timestamp BETWEEN %s AND %s 
        """
        params = [start_date, end_date]
        
        # Si el usuario NO es admin, aplicamos la restricción
        if user_role != 'admin':
            sql_log += " AND u.id = %s" 
            params.append(user_id)      
            
        sql_log += " ORDER BY f.timestamp ASC"
        
        cursor.execute(sql_log, params)
        log_records = cursor.fetchall()
        
        # ... (Formateo de fechas y cierre de DB) ...
        for record in log_records:
            if isinstance(record['timestamp'], datetime.datetime):
                record['fecha'] = record['timestamp'].strftime('%d/%m/%Y')
                record['hora'] = record['timestamp'].strftime('%H:%M:%S')
            
        if not log_records:
             f_inicio_display = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
             f_fin_display = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
             flash(f"No se encontraron fichajes entre {f_inicio_display} y {f_fin_display}.", "error")
             return redirect(url_for('dashboard'))

        # Renderiza el nuevo template, pasando los datos
        return render_template('fichajes_log.html', 
                               records=log_records, 
                               start_date=start_date_str,
                               end_date=end_date_str,
                               user_role=user_role)
                               

    except mysql.connector.Error as err:
        flash(f"Error de base de datos al buscar log: {err}", "error")
        return redirect(url_for('dashboard'))
    except Exception as e:
        flash(f"Error: {e}", "error")
        return redirect(url_for('dashboard'))
    

# --- CLASE PARA EL REPORTE PDF ---
class PDF(FPDF):
    def header(self):
        # Logo 
        # self.image('static/logo-inti.png', 10, 8, 33)
        self.set_font('Arial', 'B', 15)
        self.cell(0, 10, 'Registro de Fichajes', 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', 0, 0, 'C')


# --- RUTA PARA DESCARGAR PDF DEL LOG ---
@app.route("/admin/descargar_log_pdf", methods=['POST'])
@admin_required
def descargar_log_pdf():
    try:
        # Recibimos las fechas (vienen de campos ocultos en el HTML)
        start_date_str = request.form['log_start_date']
        end_date_str = request.form['log_end_date']
        
        start_date = f"{start_date_str} 00:00:00"
        end_date = f"{end_date_str} 23:59:59"

        # Consulta a la DB
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        sql_log = """
            SELECT f.timestamp, f.tipo, u.nombre, u.apellido, u.legajo
            FROM fichajes f
            JOIN usuarios u ON f.usuario_id = u.id
            WHERE f.timestamp BETWEEN %s AND %s
            ORDER BY f.timestamp ASC
        """
        cursor.execute(sql_log, (start_date, end_date))
        records = cursor.fetchall()
        cursor.close()
        conn.close()

        # --- CREACIÓN DEL PDF ---
        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Arial', '', 10)

        # Subtítulo con fechas
        pdf.set_font('Arial', 'B', 11)
        f_inicio = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        f_fin = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        pdf.cell(0, 10, f'Periodo: {f_inicio} al {f_fin}', 0, 1, 'L')
        pdf.ln(2)

        # Encabezados de Tabla (sin cambios)
        pdf.set_fill_color(200, 220, 255)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(30, 7, 'Fecha', 1, 0, 'C', 1)
        pdf.cell(25, 7, 'Hora', 1, 0, 'C', 1)
        pdf.cell(25, 7, 'Tipo', 1, 0, 'C', 1)
        pdf.cell(25, 7, 'Legajo', 1, 0, 'C', 1)
        pdf.cell(85, 7, 'Apellido y Nombre', 1, 1, 'C', 1)

        # Datos (sin cambios)
        pdf.set_font('Arial', '', 10)
        for row in records:
            fecha_fmt = row['timestamp'].strftime('%d/%m/%Y')
            hora_fmt = row['timestamp'].strftime('%H:%M:%S')
            nombre_completo = f"{row['apellido']}, {row['nombre']}"
            tipo = row['tipo'].upper()

            pdf.cell(30, 7, fecha_fmt, 1, 0, 'C')
            pdf.cell(25, 7, hora_fmt, 1, 0, 'C')
            
            if tipo == 'SALIDA':
                pdf.set_text_color(220, 50, 50)
            else:
                pdf.set_text_color(0, 128, 0)
            pdf.cell(25, 7, tipo, 1, 0, 'C')
            
            pdf.set_text_color(0, 0, 0)
            pdf.cell(25, 7, str(row['legajo']), 1, 0, 'C')
            pdf.cell(85, 7, nombre_completo, 1, 1, 'L')
        
        # Crear el nombre del archivo
        nombre_archivo = f"Registro primario asistencia Mar del Plata {f_inicio.replace('/', '-')}.pdf"

        # Guardar en memoria y enviar
        output = BytesIO()
        pdf_bytes = pdf.output(dest='S').encode('latin-1') 
        output.write(pdf_bytes)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre_archivo # Usamos el nuevo nombre dinámico
        )

    except Exception as e:
        print(f"Error PDF: {e}")
        flash("Error al generar PDF", "error")
        return redirect(url_for('dashboard'))
    

# --- GESTIÓN MANUAL DE FICHAJES (CRUD) ---

# 1. VER LISTA Y AGREGAR NUEVO
@app.route("/admin/fichajes", methods=['GET', 'POST'])
@admin_required
def admin_fichajes():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    
    # Fecha seleccionada (por defecto HOY)
    fecha_filtro = request.args.get('fecha', date.today().isoformat())

    # --- LOGICA PARA AGREGAR UNO NUEVO (POST) ---
    if request.method == 'POST':
        try:
            usuario_id = request.form['usuario_id']
            fecha_input = request.form['fecha']
            hora_input = request.form['hora']
            tipo = request.form['tipo']
            
            # Combinar fecha y hora en un string DATETIME
            timestamp_str = f"{fecha_input} {hora_input}"
            
            sql_insert = "INSERT INTO fichajes (usuario_id, timestamp, tipo) VALUES (%s, %s, %s)"
            cursor.execute(sql_insert, (usuario_id, timestamp_str, tipo))
            conn.commit()
            flash("Fichaje agregado manualmente con éxito.", "success")
            # Redirigir a la misma fecha para ver el cambio
            return redirect(url_for('admin_fichajes', fecha=fecha_input))
            
        except mysql.connector.Error as err:
            flash(f"Error al agregar: {err}", "error")
        
    # --- LOGICA PARA MOSTRAR LA LISTA (GET) ---
    
    # 1. Obtener usuarios para el menú desplegable (Select)
    cursor.execute("SELECT id, nombre, apellido, legajo FROM usuarios ORDER BY apellido, nombre")
    lista_usuarios = cursor.fetchall()
    
    # 2. Obtener fichajes de la fecha seleccionada
    # Buscamos desde las 00:00:00 hasta las 23:59:59 de ese día
    start_dt = f"{fecha_filtro} 00:00:00"
    end_dt = f"{fecha_filtro} 23:59:59"
    
    sql_list = """
        SELECT f.id, f.timestamp, f.tipo, u.nombre, u.apellido, u.legajo
        FROM fichajes f
        JOIN usuarios u ON f.usuario_id = u.id
        WHERE f.timestamp BETWEEN %s AND %s
        ORDER BY f.timestamp ASC
    """
    cursor.execute(sql_list, (start_dt, end_dt))
    fichajes_dia = cursor.fetchall()
    
    # Formatear hora para la vista
    for f in fichajes_dia:
        if isinstance(f['timestamp'], datetime.datetime):
            f['hora_str'] = f['timestamp'].strftime('%H:%M:%S')

    cursor.close()
    conn.close()
    
    return render_template('admin_fichajes.html', 
                           fichajes=fichajes_dia, 
                           usuarios=lista_usuarios, 
                           fecha_seleccionada=fecha_filtro)


# 2. EDITAR FICHAJE
@app.route("/admin/fichajes/editar/<int:fichaje_id>", methods=['GET', 'POST'])
@admin_required
def editar_fichaje(fichaje_id):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    
    if request.method == 'POST':
        try:
            fecha_input = request.form['fecha']
            hora_input = request.form['hora']
            tipo = request.form['tipo']
            # Nota: Generalmente no se permite cambiar el USUARIO de un fichaje, solo la hora/tipo.
            
            timestamp_str = f"{fecha_input} {hora_input}"
            
            sql_update = "UPDATE fichajes SET timestamp=%s, tipo=%s WHERE id=%s"
            cursor.execute(sql_update, (timestamp_str, tipo, fichaje_id))
            conn.commit()
            flash("Fichaje actualizado.", "success")
            return redirect(url_for('admin_fichajes', fecha=fecha_input))
            
        except mysql.connector.Error as err:
            flash(f"Error al actualizar: {err}", "error")
    
    # GET: Obtener datos actuales
    cursor.execute("""
        SELECT f.*, u.nombre, u.apellido 
        FROM fichajes f 
        JOIN usuarios u ON f.usuario_id = u.id 
        WHERE f.id = %s
    """, (fichaje_id,))
    fichaje = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not fichaje:
        flash("Fichaje no encontrado", "error")
        return redirect(url_for('admin_fichajes'))

    # Separar fecha y hora para los inputs HTML
    if isinstance(fichaje['timestamp'], datetime.datetime):
        fichaje['fecha_val'] = fichaje['timestamp'].strftime('%Y-%m-%d')
        fichaje['hora_val'] = fichaje['timestamp'].strftime('%H:%M') # Input time usa HH:MM

    return render_template('editar_fichaje.html', fichaje=fichaje)


# 3. ELIMINAR FICHAJE
@app.route("/admin/fichajes/eliminar/<int:fichaje_id>", methods=['POST'])
@admin_required
def eliminar_fichaje(fichaje_id):
    # Necesitamos saber la fecha para redirigir correctamente después de borrar
    fecha_retorno = request.form.get('fecha_retorno', date.today().isoformat())
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM fichajes WHERE id = %s", (fichaje_id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash("Fichaje eliminado correctamente.", "success")
    except mysql.connector.Error as err:
        flash(f"Error al eliminar: {err}", "error")
        
    return redirect(url_for('admin_fichajes', fecha=fecha_retorno))


# --- 4. INICIA EL SERVIDOR CON SOCKETIO ---
if __name__ == "__main__":
    print("Iniciando servidor Flask con SocketIO...")
    # Usamos socketio.run() en lugar de app.run()
    # allow_unsafe_werkzeug=True es necesario para el modo debug con versiones recientes
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, allow_unsafe_werkzeug=True)