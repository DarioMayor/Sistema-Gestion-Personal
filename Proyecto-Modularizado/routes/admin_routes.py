from flask import Blueprint, render_template, session, redirect, url_for, flash, request, send_file
import mysql.connector
from werkzeug.security import generate_password_hash
import pandas as pd
from io import BytesIO
import datetime 
from datetime import date 
import numpy as np
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import zipfile

from config import Config
from utils.decorators import login_required, admin_required
from utils.pdf_generator import PDF # Importamos la clase PDF desde utils

admin_bp = Blueprint('admin', __name__)

# --- DASHBOARD ---
@admin_bp.route("/dashboard")
def dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('auth.login'))

    contexto = {
        'username': session['username'],
        'role': session['role'],
        'nombre': session['nombre'],
        'apellido': session['apellido']
    }
    return render_template('dashboard.html', **contexto)

# --- VER FICHAJES (LOG) ---
@admin_bp.route("/ver_fichajes", methods=['POST'])
@login_required 
def ver_fichajes():
    try:
        start_date_str = request.form['log_start_date']
        end_date_str = request.form['log_end_date']
        
        start_date = f"{start_date_str} 00:00:00"
        end_date = f"{end_date_str} 23:59:59"

        conn = mysql.connector.connect(**Config.DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        # Modificamos la consulta para traer también 'horas_laborales' y el ID del usuario
        sql_log = """
            SELECT 
                f.timestamp, f.tipo, 
                u.id as uid, u.nombre, u.apellido, u.legajo, u.horas_laborales
            FROM fichajes f
            JOIN usuarios u ON f.usuario_id = u.id
            WHERE f.timestamp BETWEEN %s AND %s 
        """
        params = [start_date, end_date]
        
        if user_role != 'admin':
            sql_log += " AND u.id = %s" 
            params.append(user_id)      
            
        sql_log += " ORDER BY f.timestamp ASC"
        
        cursor.execute(sql_log, params)
        log_records = cursor.fetchall()
        
        cursor.close()
        conn.close()

        # --- CÁLCULO DE ALERTAS ---
        # 1. Agrupar eventos por (Usuario, Fecha)
        eventos_por_dia = {}
        
        for r in log_records:
            # Convertimos a objetos Python puros para trabajar
            if isinstance(r['timestamp'], datetime.datetime):
                fecha_obj = r['timestamp'].date()
                uid = r['uid']
                clave = (uid, fecha_obj)
                
                if clave not in eventos_por_dia:
                    eventos_por_dia[clave] = {
                        'eventos': [], 
                        'horas_laborales': r['horas_laborales']
                    }
                eventos_por_dia[clave]['eventos'].append(r)

        # 2. Calcular tiempo trabajado por día y detectar deficiencias
        claves_con_alerta = set() # Guardaremos los pares (uid, fecha) que fallaron
        
        for clave, datos in eventos_por_dia.items():
            eventos = datos['eventos']
            
            # Detectar fichajes sin cierre (entrada sin salida)
            entrada_pendiente = None
            for e in eventos:
                # Inicializar alerta_amarilla en False para todos
                e['alerta_amarilla'] = False
                
                if e['tipo'] == 'entrada':
                    if entrada_pendiente:
                        # La entrada anterior no tuvo salida
                        entrada_pendiente['alerta_amarilla'] = True
                    entrada_pendiente = e
                elif e['tipo'] == 'salida':
                    if entrada_pendiente:
                        # Cierre correcto
                        entrada_pendiente = None
                    else:
                        # Salida huérfana (opcional: marcar también si se desea)
                        pass
            
            # Si quedó una entrada pendiente al final del día
            if entrada_pendiente:
                entrada_pendiente['alerta_amarilla'] = True

            horas_meta = datos['horas_laborales'] # timedelta o string
            
            # Convertir meta a segundos
            if isinstance(horas_meta, datetime.timedelta):
                segundos_meta = horas_meta.total_seconds()
            else:
                # Fallback por si viene como string "08:00:00"
                h, m, s = map(int, str(horas_meta).split(':'))
                segundos_meta = h * 3600 + m * 60 + s

            segundos_trabajados = 0
            entrada_temp = None
            
            # Sumar intervalos Entrada -> Salida
            for e in eventos:
                ts = e['timestamp']
                if e['tipo'] == 'entrada':
                    entrada_temp = ts
                elif e['tipo'] == 'salida' and entrada_temp:
                    delta = ts - entrada_temp
                    segundos_trabajados += delta.total_seconds()
                    entrada_temp = None
            
            # Si hay una entrada sin salida al final del día, no podemos calcular bien,
            # así que por ahora ignoramos ese último tramo o asumimos que sigue trabajando.
            
            # REGLA: Si trabajó menos de (Meta - 30 min)
            if segundos_trabajados < (segundos_meta - 1800):
                claves_con_alerta.add(clave)

        # 3. Marcar los registros originales y formatear para la vista
        for record in log_records:
            # Formato de fecha/hora
            if isinstance(record['timestamp'], datetime.datetime):
                fecha_obj = record['timestamp'].date()
                record['fecha'] = record['timestamp'].strftime('%d/%m/%Y')
                record['hora'] = record['timestamp'].strftime('%H:%M:%S')
                
                # Verificar alerta
                uid = record['uid']
                if (uid, fecha_obj) in claves_con_alerta:
                    record['alerta'] = True
                else:
                    record['alerta'] = False
                
                # La alerta amarilla ya se calculó individualmente en el paso anterior
                if 'alerta_amarilla' not in record:
                    record['alerta_amarilla'] = False
            
        if not log_records:
             f_inicio_display = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
             f_fin_display = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
             flash(f"No se encontraron fichajes entre {f_inicio_display} y {f_fin_display}.", "error")
             return redirect(url_for('admin.dashboard'))

        return render_template('fichajes_log.html', 
                               records=log_records, 
                               start_date=start_date_str,
                               end_date=end_date_str,
                               user_role=user_role)

    except mysql.connector.Error as err:
        flash(f"Error de base de datos al buscar log: {err}", "error")
        return redirect(url_for('admin.dashboard'))
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Error: {e}", "error")
        return redirect(url_for('admin.dashboard'))
    
    
# --- GESTIÓN USUARIOS (LISTA) ---
@admin_bp.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    try:
        conn = mysql.connector.connect(**Config.DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, nombre, apellido, legajo, username, role FROM usuarios ORDER BY apellido, nombre")
        lista_de_usuarios = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('admin_usuarios.html', usuarios=lista_de_usuarios)
    except mysql.connector.Error as err:
        flash(f"Error al cargar usuarios: {err}", "error")
        return redirect(url_for('admin.dashboard'))

# --- CREAR USUARIO ---
@admin_bp.route("/admin/crear", methods=['GET', 'POST'])
@admin_required
def crear_usuario():
    form_data = request.form if request.method == 'POST' else {}

    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        legajo = request.form['legajo']
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        huellas_str = request.form.get('huellas', '')
        horas_laborales = request.form['horas_laborales']
        
        password_hash = generate_password_hash(password)

        conn = mysql.connector.connect(**Config.DB_CONFIG)
        cursor = conn.cursor()
        
        try:
            sql_insert_user = """
                INSERT INTO usuarios (nombre, apellido, legajo, horas_laborales, username, password_hash, role) 
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql_insert_user, (nombre, apellido, legajo, horas_laborales, username, password_hash, role))
            new_user_id = cursor.lastrowid
            cursor.execute(sql_insert_user, (nombre, apellido, legajo, username, password_hash, role))
            new_user_id = cursor.lastrowid
            
            if new_user_id and huellas_str:
                huellas_list = [h.strip() for h in huellas_str.split(',') if h.strip().isdigit()]
                if huellas_list:
                    sql_insert_huella = "INSERT INTO huellas (huella_id, usuario_id) VALUES (%s, %s)"
                    for huella_id in set(huellas_list):
                        cursor.execute(sql_insert_huella, (huella_id, new_user_id))
            
            conn.commit()
            flash(f"Usuario '{username}' creado exitosamente.", "success")
            return redirect(url_for('admin.admin_usuarios'))
        
        except mysql.connector.Error as err:
            conn.rollback()
            if err.errno == 1062: 
                flash(f"Error: El Email, Legajo o una de las Huellas ya está en uso.", "error")
            else:
                flash(f"Error de base de datos: {err}", "error")
            return render_template('crear_usuario.html', form_data=form_data, huellas_str=huellas_str)
        
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    return render_template('crear_usuario.html', form_data=form_data)

# --- EDITAR USUARIO ---
@admin_bp.route("/admin/editar/<int:usuario_id>", methods=['GET', 'POST'])
@admin_required
def editar_usuario(usuario_id):
    conn = mysql.connector.connect(**Config.DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    
    usuario_a_editar = None
    huellas_str = ''
    
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        legajo = request.form['legajo']
        username = request.form['username']
        role = request.form['role']
        nueva_password = request.form['nueva_password']
        huellas_str = request.form.get('huellas', '')
        horas_laborales = request.form['horas_laborales']
        
        try:
            if nueva_password:
                password_hash = generate_password_hash(nueva_password)
                sql_update = "UPDATE usuarios SET nombre=%s, apellido=%s, legajo=%s, username=%s, role=%s, password_hash=%s, horas_laborales=%s WHERE id=%s"
                cursor.execute(sql_update, (nombre, apellido, legajo, username, role, password_hash, horas_laborales, usuario_id))
            else:
                sql_update = "UPDATE usuarios SET nombre=%s, apellido=%s, legajo=%s, username=%s, role=%s, horas_laborales=%s WHERE id=%s"
                cursor.execute(sql_update, (nombre, apellido, legajo, username, role, usuario_id, horas_laborales))
            
            cursor.execute("DELETE FROM huellas WHERE usuario_id = %s", (usuario_id,))
            if huellas_str:
                huellas_list = [h.strip() for h in huellas_str.split(',') if h.strip().isdigit()]
                if huellas_list:
                    sql_insert_huella = "INSERT INTO huellas (huella_id, usuario_id) VALUES (%s, %s)"
                    for huella_id in set(huellas_list):
                        cursor.execute(sql_insert_huella, (huella_id, usuario_id))
            
            conn.commit()
            flash(f"Usuario '{username}' actualizado exitosamente.", "success")
            return redirect(url_for('admin.admin_usuarios'))
        
        except mysql.connector.Error as err:
            conn.rollback() 
            if err.errno == 1062:
                flash(f"Error: Datos duplicados (Email, Legajo o Huella).", "error")
            else:
                flash(f"Error de base de datos: {err}", "error")
            
            usuario_a_editar = { 'id': usuario_id, 'nombre': nombre, 'apellido': apellido, 'legajo': legajo, 'username': username, 'role': role, 'horas_laborales': horas_laborales }
            return render_template('editar_usuario.html', usuario=usuario_a_editar, huellas_str=huellas_str)
        
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    try:
        cursor.execute("SELECT id, nombre, apellido, legajo, username, role, horas_laborales FROM usuarios WHERE id = %s", (usuario_id,))
        usuario_a_editar = cursor.fetchone()
        cursor.execute("SELECT huella_id FROM huellas WHERE usuario_id = %s", (usuario_id,))
        huellas_raw = cursor.fetchall()
        huellas_str = ', '.join([str(h['huella_id']) for h in huellas_raw])
        
        if not usuario_a_editar:
            flash("Error: Usuario no encontrado.", "error")
            return redirect(url_for('admin.admin_usuarios'))
            
        return render_template('editar_usuario.html', usuario=usuario_a_editar, huellas_str=huellas_str)
    
    except mysql.connector.Error as err:
        flash(f"Error de base de datos: {err}", "error")
        return redirect(url_for('admin.admin_usuarios'))
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# --- ELIMINAR USUARIO ---
@admin_bp.route("/admin/eliminar/<int:usuario_id>", methods=['POST'])
@admin_required
def eliminar_usuario(usuario_id):
    if usuario_id == session['user_id']:
        flash("No puedes eliminar tu propia cuenta.", "error")
        return redirect(url_for('admin.admin_usuarios'))
    try:
        conn = mysql.connector.connect(**Config.DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM huellas WHERE usuario_id = %s", (usuario_id,))
        cursor.execute("DELETE FROM fichajes WHERE usuario_id = %s", (usuario_id,))
        cursor.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
        conn.commit()
        flash("Usuario eliminado exitosamente.", "success")
    except mysql.connector.Error as err:
        flash(f"Error al eliminar: {err}", "error")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
    return redirect(url_for('admin.admin_usuarios'))

# --- DESCARGAR EXCEL ---
@admin_bp.route("/admin/descargar_excel", methods=['POST'])
@admin_required
def descargar_excel():
    try:
        start_date_str = request.form['start_date']
        end_date_str = request.form['end_date']
        start_date = f"{start_date_str} 00:00:00"
        end_date = f"{end_date_str} 23:59:59"
        
        # Generar rango de fechas COMPLETO para el reporte
        dt_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        dt_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        # pd.date_range crea todas las fechas intermedias
        all_dates = pd.date_range(start=dt_start, end=dt_end).date

        try:
            start_date_obj = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
            meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
            month_name = meses[start_date_obj.month].capitalize()
        except:
            month_name = ""

        conn = mysql.connector.connect(**Config.DB_CONFIG)
        sql_query = """SELECT u.legajo, u.nombre, u.apellido, f.timestamp, f.tipo FROM fichajes f JOIN usuarios u ON f.usuario_id = u.id WHERE f.timestamp BETWEEN %s AND %s ORDER BY u.legajo, f.timestamp"""
        df = pd.read_sql(sql_query, conn, params=(start_date, end_date))
        conn.close()

        if df.empty:
            flash("No se encontraron fichajes.", "error")
            return redirect(url_for('admin.dashboard'))

        df['fecha'] = df['timestamp'].dt.date
        df['hora'] = df['timestamp'].dt.time
        
        entradas_df = df[df['tipo'] == 'entrada']
        salidas_df = df[df['tipo'] == 'salida']
        common_cols = ['legajo', 'nombre', 'apellido', 'fecha']
        
        e_min = entradas_df.groupby(common_cols)['hora'].min().reset_index().rename(columns={'hora': 'Entrada'})
        s_min = salidas_df.groupby(common_cols)['hora'].min().reset_index().rename(columns={'hora': 'Salida'})
        e_max = entradas_df.groupby(common_cols)['hora'].max().reset_index().rename(columns={'hora': 'Entrada'})
        s_max = salidas_df.groupby(common_cols)['hora'].max().reset_index().rename(columns={'hora': 'Salida'})

        # Limpieza duplicados
        e_max = pd.merge(e_max, e_min, on=common_cols, suffixes=('', '_min'), how='left')
        mask_e = e_max['Entrada'] == e_max['Entrada_min']
        e_max.loc[mask_e, 'Entrada'] = None 
        e_max = e_max.drop(columns=['Entrada_min'])
        
        s_max = pd.merge(s_max, s_min, on=common_cols, suffixes=('', '_min'), how='left')
        mask_s = s_max['Salida'] == s_max['Salida_min']
        s_max.loc[mask_s, 'Salida'] = None 
        s_max = s_max.drop(columns=['Salida_min'])

        e_min['Turno'] = 'Mañana'; e_max['Turno'] = 'Tarde'
        s_min['Turno'] = 'Mañana'; s_max['Turno'] = 'Tarde'

        all_entradas = pd.concat([e_min, e_max])
        all_salidas = pd.concat([s_min, s_max])
        merge_cols = ['legajo', 'nombre', 'apellido', 'fecha', 'Turno']
        reporte_long_df = pd.merge(all_entradas, all_salidas, on=merge_cols, how='outer')
        reporte_long_df = reporte_long_df.drop_duplicates(subset=['legajo', 'fecha', 'Entrada', 'Salida'])
        reporte_long_df['Apellido y Nombre'] = reporte_long_df['apellido'].fillna('') + ', ' + reporte_long_df['nombre'].fillna('')
        
        for col in ['Entrada', 'Salida']:
            reporte_long_df[col] = reporte_long_df[col].apply(lambda x: x.strftime('%H:%M') if isinstance(x, datetime.time) else '')

        pivot_df = reporte_long_df.pivot_table(index=['legajo', 'Apellido y Nombre', 'Turno'], columns='fecha', values=['Entrada', 'Salida'], aggfunc='first', fill_value='')
        
        if pivot_df.empty:
             flash("No hay datos para generar el reporte.", "error")
             return redirect(url_for('admin.dashboard'))

        pivot_df = pivot_df.swaplevel(0, 1, axis=1).sort_index(axis=1)
        
        # --- AQUÍ ESTÁ LA MAGIA: REINDEXAR CON TODAS LAS FECHAS ---
        # Creamos un MultiIndex con TODAS las fechas y las columnas Entrada/Salida
        full_columns = pd.MultiIndex.from_product([all_dates, ['Entrada', 'Salida']], names=['fecha', 'tipo'])
        # Reindexamos para forzar que aparezcan todas las fechas, incluso las vacías
        pivot_df = pivot_df.reindex(columns=full_columns, fill_value='')
        # ----------------------------------------------------------

        pivot_df = pivot_df.reindex(pd.Categorical(['Mañana', 'Tarde'], ordered=True), level='Turno')
        pivot_df = pivot_df.sort_index(level=['legajo', 'Turno'])
        pivot_df = pivot_df.reset_index()
        
        # Aplanar MultiIndex
        new_cols = []
        for col in pivot_df.columns:
            if isinstance(col, tuple):
                new_cols.append('_'.join(map(str, col)).strip('_'))
            else:
                new_cols.append(str(col))
        pivot_df.columns = new_cols

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        pd.DataFrame().to_excel(writer, sheet_name='Reporte Fichajes', header=False, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Reporte Fichajes']
        # Definir Estilos
        bold_font = Font(bold=True)
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        medium_side = Side(style='medium'); thin_side = Side(style='thin'); gray_thin_side = Side(style='thin', color="A9A9A9")
        border_R1 = Border(top=medium_side, left=medium_side, right=medium_side, bottom=thin_side)
        border_A2 = Border(top=thin_side, left=medium_side, bottom=medium_side, right=thin_side)
        border_B2 = Border(top=thin_side, left=thin_side, bottom=medium_side, right=medium_side)
        border_R2_entrada = Border(top=thin_side, left=medium_side, bottom=medium_side, right=thin_side)
        border_R2_salida = Border(top=thin_side, left=thin_side, bottom=medium_side, right=medium_side)
        border_legajo_top = Border(left=medium_side, top=medium_side, right=thin_side, bottom=thin_side)
        border_legajo_bottom = Border(left=medium_side, top=thin_side, right=thin_side, bottom=medium_side)
        border_nombre_top = Border(left=thin_side, top=medium_side, right=medium_side, bottom=thin_side)
        border_nombre_bottom = Border(left=thin_side, top=thin_side, right=medium_side, bottom=medium_side)
        border_data_entrada_am = Border(left=medium_side, top=medium_side, right=gray_thin_side, bottom=gray_thin_side)
        border_data_salida_am = Border(left=gray_thin_side, top=medium_side, right=medium_side, bottom=gray_thin_side)
        border_data_entrada_pm = Border(left=medium_side, top=gray_thin_side, right=gray_thin_side, bottom=medium_side)
        border_data_salida_pm = Border(left=gray_thin_side, top=gray_thin_side, right=medium_side, bottom=medium_side)

        cell_a1 = worksheet['A1']; cell_a1.value = f"Mar del Plata - {month_name}"; cell_a1.font = bold_font; cell_a1.fill = silver_fill; cell_a1.border = border_R1; cell_a1.alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:B1'); cell_b1 = worksheet['B1']; cell_b1.font = bold_font; cell_b1.fill = silver_fill; cell_b1.border = border_R1; cell_b1.alignment = Alignment(horizontal='center')
        
        # Usamos all_dates en lugar de fechas del DF para el bucle de encabezados
        fechas_reales = sorted(all_dates)
        
        col_idx = 3
        for fecha in fechas_reales:
            worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
            cell = worksheet.cell(row=1, column=col_idx); cell.value = fecha.strftime('%d/%m/%Y'); cell.alignment = Alignment(horizontal='center'); cell.font = bold_font; cell.fill = silver_fill; cell.border = border_R1
            worksheet.cell(row=1, column=col_idx + 1).font = bold_font; worksheet.cell(row=1, column=col_idx + 1).fill = silver_fill; worksheet.cell(row=1, column=col_idx + 1).border = border_R1
            col_idx += 2
        cell_a2 = worksheet['A2']; cell_a2.value = "Legajo"; cell_a2.font = bold_font; cell_a2.fill = silver_fill; cell_a2.border = border_A2; cell_a2.alignment = Alignment(horizontal='center')
        cell_b2 = worksheet['B2']; cell_b2.value = "Apellido y Nombre"; cell_b2.font = bold_font; cell_b2.fill = silver_fill; cell_b2.border = border_B2; cell_b2.alignment = Alignment(horizontal='center')
        col_idx = 3
        for fecha in fechas_reales:
            cell_e = worksheet.cell(row=2, column=col_idx); cell_e.value = "Entrada"; cell_e.font = bold_font; cell_e.fill = silver_fill; cell_e.border = border_R2_entrada; cell_e.alignment = Alignment(horizontal='center')
            cell_s = worksheet.cell(row=2, column=col_idx + 1); cell_s.value = "Salida"; cell_s.font = bold_font; cell_s.fill = silver_fill; cell_s.border = border_R2_salida; cell_s.alignment = Alignment(horizontal='center')
            col_idx += 2
        current_excel_row = 3
        
        # Agrupar por las columnas que ahora son strings tras aplanar
        for (legajo, apellido_nombre), group in pivot_df.groupby(['legajo', 'Apellido y Nombre']):
            row_am = group[group['Turno'] == 'Mañana']; row_pm = group[group['Turno'] == 'Tarde']
            
            cell_legajo = worksheet.cell(row=current_excel_row, column=1, value=str(legajo)); cell_legajo.border = border_legajo_top; cell_legajo.alignment = Alignment(vertical='center', horizontal='center')
            worksheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row + 1, end_column=1)
            worksheet.cell(row=current_excel_row+1, column=1).border = border_legajo_bottom
            cell_nombre = worksheet.cell(row=current_excel_row, column=2, value=apellido_nombre); cell_nombre.border = border_nombre_top; cell_nombre.alignment = Alignment(vertical='center')
            worksheet.merge_cells(start_row=current_excel_row, start_column=2, end_row=current_excel_row + 1, end_column=2)
            worksheet.cell(row=current_excel_row+1, column=2).border = border_nombre_bottom
            col_idx = 3
            for fecha in fechas_reales:
                # Reconstruir el nombre de columna aplanado
                fecha_str = str(fecha)
                col_ent = f"{fecha_str}_Entrada"
                col_sal = f"{fecha_str}_Salida"
                
                val_e = row_am[col_ent].values[0] if not row_am.empty and col_ent in row_am else ''; cell_e_am = worksheet.cell(row=current_excel_row, column=col_idx, value=val_e); cell_e_am.border = border_data_entrada_am; cell_e_am.alignment = Alignment(horizontal='center')
                val_s = row_am[col_sal].values[0] if not row_am.empty and col_sal in row_am else ''; cell_s_am = worksheet.cell(row=current_excel_row, column=col_idx + 1, value=val_s); cell_s_am.border = border_data_salida_am; cell_s_am.alignment = Alignment(horizontal='center')
                col_idx += 2
            current_excel_row += 1
            col_idx = 3
            for fecha in fechas_reales:
                fecha_str = str(fecha)
                col_ent = f"{fecha_str}_Entrada"
                col_sal = f"{fecha_str}_Salida"
                val_e = row_pm[col_ent].values[0] if not row_pm.empty and col_ent in row_pm else ''; cell_e_pm = worksheet.cell(row=current_excel_row, column=col_idx, value=val_e); cell_e_pm.border = border_data_entrada_pm; cell_e_pm.alignment = Alignment(horizontal='center')
                val_s = row_pm[col_sal].values[0] if not row_pm.empty and col_sal in row_pm else ''; cell_s_pm = worksheet.cell(row=current_excel_row, column=col_idx + 1, value=val_s); cell_s_pm.border = border_data_salida_pm; cell_s_pm.alignment = Alignment(horizontal='center')
                col_idx += 2
            current_excel_row += 1
        worksheet.column_dimensions['B'].width = 30; worksheet.column_dimensions['A'].width = 15

        writer.close(); output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='reporte_fichaje.xlsx')

    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Error al generar el reporte: {e}", "error")
        return redirect(url_for('admin.dashboard'))

# --- DESCARGAR PDF ---
@admin_bp.route("/admin/descargar_log_pdf", methods=['POST'])
@login_required
def descargar_log_pdf():
    try:
        start_date_str = request.form['log_start_date']
        end_date_str = request.form['log_end_date']
        
        # Convertir strings a objetos date
        d_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        d_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Generar lista de días en el rango
        delta = d_end - d_start
        lista_dias = [d_start + datetime.timedelta(days=i) for i in range(delta.days + 1)]
        
        user_role = session.get('role')
        user_id = session.get('user_id')
        
        conn = mysql.connector.connect(**Config.DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        
        # Diccionario para guardar los PDFs en memoria: { "nombre_archivo.pdf": bytes_pdf }
        archivos_pdf = {}

        for dia in lista_dias:
            dia_str = dia.strftime('%Y-%m-%d')
            start_limit = f"{dia_str} 00:00:00"
            end_limit = f"{dia_str} 23:59:59"
            
            # Consulta SOLO para este día
            sql_log = """
                SELECT f.timestamp, f.tipo, u.nombre, u.apellido, u.legajo
                FROM fichajes f
                JOIN usuarios u ON f.usuario_id = u.id
                WHERE f.timestamp BETWEEN %s AND %s
            """
            params = [start_limit, end_limit]
            
            if user_role != 'admin':
                sql_log += " AND u.id = %s" 
                params.append(user_id)
                
            sql_log += " ORDER BY f.timestamp ASC"
            
            cursor.execute(sql_log, params)
            records = cursor.fetchall()
            
            # Si no hay registros ese día, podemos saltarlo o generar PDF vacío.
            # Vamos a saltarlo para no llenar el ZIP de archivos vacíos.
            if not records:
                continue

            # --- Generar PDF para este día ---
            pdf = PDF()
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.set_font('Arial', '', 10)
            
            # Título del día
            fecha_fmt_titulo = dia.strftime('%d/%m/%Y')
            pdf.cell(0, 10, f'Fecha: {fecha_fmt_titulo}', 0, 1, 'L')
            pdf.ln(2)

            pdf.set_fill_color(200, 220, 255)
            pdf.set_font('Arial', 'B', 10)
            
            w_fecha, w_hora, w_tipo, w_legajo, w_nombre = 30, 25, 25, 25, 85
            if user_role == 'admin':
                pdf.cell(w_fecha, 7, 'Fecha', 1, 0, 'C', 1)
                pdf.cell(w_hora, 7, 'Hora', 1, 0, 'C', 1)
                pdf.cell(w_tipo, 7, 'Tipo', 1, 0, 'C', 1)
                pdf.cell(w_legajo, 7, 'Legajo', 1, 0, 'C', 1)
                pdf.cell(w_nombre, 7, 'Apellido y Nombre', 1, 1, 'C', 1)
            else:
                pdf.cell(w_fecha, 7, 'Fecha', 1, 0, 'C', 1)
                pdf.cell(w_hora, 7, 'Hora', 1, 0, 'C', 1)
                pdf.cell(w_tipo, 7, 'Tipo', 1, 1, 'C', 1)

            pdf.set_font('Arial', '', 10)
            for row in records:
                fecha_fmt = row['timestamp'].strftime('%d/%m/%Y')
                hora_fmt = row['timestamp'].strftime('%H:%M:%S')
                nombre_completo = f"{row['apellido']}, {row['nombre']}"
                tipo = row['tipo'].upper()

                pdf.cell(w_fecha, 7, fecha_fmt, 1, 0, 'C')
                pdf.cell(w_hora, 7, hora_fmt, 1, 0, 'C')
                
                if tipo == 'SALIDA': pdf.set_text_color(220, 50, 50)
                else: pdf.set_text_color(0, 128, 0)
                pdf.cell(w_tipo, 7, tipo, 1, 0, 'C')
                pdf.set_text_color(0, 0, 0)
                
                if user_role == 'admin':
                    pdf.cell(w_legajo, 7, str(row['legajo']), 1, 0, 'C')
                    pdf.cell(w_nombre, 7, nombre_completo, 1, 1, 'L')
                else:
                    pdf.cell(0, 7, '', 0, 1, 'L') 

            # Guardar PDF en memoria
            pdf_bytes = pdf.output(dest='S').encode('latin-1')
            nombre_archivo_dia = f"Registro primario asistencia Mar del Plata {dia.strftime('%d-%m-%Y')}.pdf"
            archivos_pdf[nombre_archivo_dia] = pdf_bytes

        cursor.close()
        conn.close()
        
        if not archivos_pdf:
            flash("No se encontraron registros en los días seleccionados.", "error")
            return redirect(url_for('admin.dashboard'))
            
        if len(archivos_pdf) == 1:
            # Solo un archivo: Descargar PDF directo
            nombre, contenido = list(archivos_pdf.items())[0]
            output = BytesIO(contenido)
            return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=nombre)
            
        else:
            # Varios archivos: Crear ZIP
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for nombre, contenido in archivos_pdf.items():
                    zip_file.writestr(nombre, contenido)
            
            zip_buffer.seek(0)
            nombre_zip = f"Registros_Fichajes_{start_date_str}_al_{end_date_str}.zip"
            
            return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=nombre_zip)

    except Exception as e:
        print(f"Error PDF/ZIP: {e}")
        import traceback; traceback.print_exc()
        flash("Error al generar archivos", "error")
        return redirect(url_for('admin.dashboard'))